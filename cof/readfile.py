from openpyxl import  load_workbook

def readfile(filename,sheetname):
    wb=load_workbook(filename)
    sheetobj=wb[sheetname]
    Excel_list=[]
    for i in range(2,sheetobj.max_row+1):
        list=[]
        for j in range(1,sheetobj.max_column+1):
            list.append(sheetobj.cell(i,j).value)
        Excel_list.append(list)
    return  Excel_list
